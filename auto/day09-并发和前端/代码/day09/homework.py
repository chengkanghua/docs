import os
import time
import configparser

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SETTINGS_PATH = os.path.join(BASE_DIR, 'settings.ini')

MESSAGE_DICT = {
    "447472516": "上一次发送报警的时间"
}


class SettingsContext:
    def __init__(self):
        self.wx_url = str
        self.db_path = str


def load_settings():
    try:

        """ 加载配置文件：读取本地的ini文件 """
        if not os.path.exists(SETTINGS_PATH):
            input("[异常]配置文件settings.ini不存在")
            return

        # 读取ini文件的内容
        config_parser = configparser.ConfigParser()
        config_parser.read(SETTINGS_PATH, encoding='utf-8')

        obj = SettingsContext()
        for name, type_cls in obj.__dict__.items():
            value = config_parser.get('settings', name)
            setattr(obj, name, type_cls(value))
        return obj
    except Exception as e:
        input("[异常]加载settings.ini配置错误，详细信息：{}".format(str(e)))


def load_excel(file_path):
    if not os.path.exists(file_path):
        input("[异常]excel文件不存在")
        return

    # 读取所有excel文件将所有的型号全都放在列表中返回。
    model_list = []
    wb = load_workbook(file_path)
    sheet = wb.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        model = row[0].value
        if not model:
            continue
        model_list.append(model)
    return model_list


def send_message(web_hook_url, content):
    res = requests.post(
        url=web_hook_url,
        json={
            "msgtype": "text",
            "text": {
                "content": content,
                "mentioned_list": ["@all"]
            }
        }

    )
    res.close()


def do_search(model, wx_url):
    """
    进行搜索
    :param model:型号
    :param wx_url:微信通知地址
    """
    try:
        url = 'http://mall.10010.com/goodsdetail/{}.html'.format(model)
        res = requests.get(url)
        soup = BeautifulSoup(res.text, features='html.parser')
        tag = soup.find(name='span', attrs={'id': "amountChange_id"}).text
        if tag != "有货":
            return

        # 有货，应该进行通知（不要频繁通知，10分钟一次）
        ctime = time.time()
        pre_send_time = MESSAGE_DICT.get(model)
        if not pre_send_time or (pre_send_time + 600) < ctime:
            send_message(wx_url, "型号：{} 有货了".format(model))
            MESSAGE_DICT[model] = ctime
            return

    except Exception as e:
        pass


def run():
    # 1.加载配置文件
    settings_object = load_settings()
    if not settings_object:
        return

    # 2.加载excel文件
    model_list = load_excel(settings_object.db_path)
    if not model_list:
        return

    # 3.开始监测
    while True:
        for model in model_list:
            do_search(model, settings_object.wx_url)


if __name__ == '__main__':
    run()
