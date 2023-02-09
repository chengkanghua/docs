from rest_framework.views import APIView

from rest_framework.response import Response

from rest_framework.permissions import IsAuthenticated
from rest_framework_jwt.authentication import JSONWebTokenAuthentication

# class HostView(APIView):
#     authentication_classes = [JSONWebTokenAuthentication]
#     permission_classes = [IsAuthenticated]
#
#     def get(self, reqeust):
#         return Response("OK")

from rest_framework.generics import ListAPIView, CreateAPIView, ListCreateAPIView
from rest_framework.viewsets import ModelViewSet
from .models import HostCategory, Host
from .serializers import HostCategoryModelSeiralizer, HostModelSerializers
from rest_framework.permissions import IsAuthenticated


class HostCategoryListAPIView(ListCreateAPIView):
    """主机类别"""
    queryset = HostCategory.objects.filter(is_show=True, is_deleted=False).order_by("orders", "-id").all()
    serializer_class = HostCategoryModelSeiralizer
    permission_classes = [IsAuthenticated]


class HostModelViewSet(ModelViewSet):
    """主机信息"""
    queryset = Host.objects.all()
    serializer_class = HostModelSerializers
    permission_classes = [IsAuthenticated]


import os
from io import BytesIO
from openpyxl import load_workbook
from .models import HostCategory


class HostExcelView(APIView):
    def post(self, request):
        host_excel = request.FILES.get("host_excel")
        print("host_excel:", host_excel, type(host_excel))
        print(host_excel.name)
        # 将文件写入到服务器
        # path = os.path.join("upload_files", host_excel.name)
        # with open(path, "wb") as f_write:
        #     for line in host_excel:
        #         f_write.write(line)

        io_data = BytesIO()
        for line in host_excel:
            io_data.write(line)

        # 加载excel文件
        wb = load_workbook(io_data)
        # 获取worksheet
        worksheet = wb.worksheets[1]

        # 获取主机类别数据
        category_list = HostCategory.objects.all().values_list("id", "name")
        print("category_list", category_list)

        host_info_list = []
        for row in worksheet.iter_rows(2):
            one_row_dict = {}  # 单个主机信息字典
            if not row[0].value: continue

            # 处理类别

            for category in category_list:
                if category[1] == row[0].value:
                    one_row_dict["category"] = category[0]
                    break

            one_row_dict["name"] = row[1].value
            one_row_dict["ip_addr"] = row[2].value
            one_row_dict["port"] = row[3].value
            one_row_dict["username"] = row[4].value

            excel_pwd = row[5].value
            try:
                pwd = str(excel_pwd)  # 这样强转容易报错，最好捕获一下异常，并记录单元格位置，给用户保存信息时，可以提示用户哪个单元格的数据有问题
            except Exception as e:
                pwd = ""
            one_row_dict["password"] = pwd

            one_row_dict["description"] = row[6].value

            print("one_row_dict", one_row_dict)
            host_info_list.append(one_row_dict)

        # 基于ORM批量创建

        # 校验主机数据
        # 将做好的主机信息字典数据通过我们添加主机时的序列化器进行校验

        res_data = {}  # 存放上传成功之后需要返回的主机数据和某些错误信息数据
        serializers_host_res_data = []
        res_error_data = []
        for k, host_data in enumerate(host_info_list):
            # 反序列化校验每一个主机信息
            serailizer = HostModelSerializers(data=host_data)
            if serailizer.is_valid():
                new_host_obj = serailizer.save()
                serializers_host_res_data.append(new_host_obj)
            else:
                # 报错，并且错误信息中应该体验错误的数据位置
                res_error_data.append({'error': f'该{k + 1}行数据有误,其他没有问题的数据，已经添加成功了，请求失败数据改完之后，重新上传这个错误数据，成功的数据不需要上传了'})

        # 再次调用序列化器进行数据的序列化，返回给客户端
        serializer = HostModelSerializers(instance=serializers_host_res_data, many=True)
        res_data['data'] = serializer.data
        res_data['error'] = res_error_data

        return Response(res_data)
