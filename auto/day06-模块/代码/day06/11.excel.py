from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell, Cell

wb = load_workbook("files/new.xlsx")
sheet = wb.worksheets[0]

# 获取第N行第N列的单元格(位置是从1开始）
c1 = sheet.cell(1, 2)
if type(c1) == MergedCell:
    print("被合并的单元格")
else:
    print("正常单元格")
